#Librerias
from openpyxl import load_workbook
import math

#Importacion de log y de la hoja
wb = load_workbook(filename="log.xlsx")
ws = wb.active

#Variables
antminimos = 10000
auxant = 100000


sumaminimos = 0
ant = 0.1
flag = 1
cont = 0

#Numero de centros inicial
ncentros = 2

an = []
pn = []
eu = [[]for _ in range(ncentros)] #Distanias euclidianas
aux = []
mejores = []

minimos = [] #Valores minimos de las distancias euclidianas
bins = [[]for _ in range(ncentros)]


#Posicion de centroides
centro = [[0,0],[1,1],[0.5,0.5],[0.3,0.3],[0.7,0.7]]
centron = [[]for _ in range(ncentros)]

#Funciones
def sumprod(x,y,n):
    sp = 0
    for i in range(n):
        sp = sp + (x[i]*y[i])
    return sp


def sumalista(x):
    suma = 0
    for i in x:
        suma = suma + i
    return suma


def impresiones():
    print("Total de filas = "+str(len(filas)))

    print("Altura = "+str(altura))
    print("Total altura = "+str(len(altura)))

    print("Peso = "+str(peso))
    print("Total altura = "+str(len(peso)))

    print("Altura minima = "+str(alturamin))
    print("Altura maxima = "+str(alturamax))
    print("Peso minimo = "+str(pesomin))
    print("Peso maximo = "+str(pesomax))

    print("Pendiente de altura = "+str(maltura))
    print("Pendiente de peso = "+str(mpeso))

    print("B Altura = "+str(baltura))
    print("B Peso = "+str(bpeso))

    print("Altura normalizada= "+str(an))
    print("Peso normalizado= "+str(pn))

#Obtener las filas del log en una matriz

filas = [[cell.value for cell in row] for row in ws.iter_rows(row_offset=1)]
del filas[len(filas)-1]

#Se vacian los datos de la matriz en arreglos separados para un control mejor sobre los datos
altura = [c[0] for c in filas]
peso = [c[1] for c in filas]

#Se sacan los minimos y maximos
alturamin = min(altura)
alturamax = max(altura)
pesomin = min(peso)
pesomax = max(peso)

#Sacar la pendiente
maltura = 1/(float(alturamax) - float(alturamin))
mpeso = 1/(float(pesomax) - float(pesomin))

#Sacar la b
baltura = -(maltura)*alturamin
bpeso = -(mpeso)*pesomin

#Normalizar la altura
for i in altura:
    an.append((maltura*i)+baltura)

#Normalizar el peso
for i in peso:
    pn.append((mpeso*i)+bpeso)


#Ciclo dinamico principal
while(flag==1):

    print(str(cont)+" VEZ")
    print(str(ncentros)+" centros")

    if(c >=1):
        antminimos = auxant

    cont = cont + 1

    #Bucle
    for iteraciones in range(10):

        #Bandera
        if iteraciones >=1:
            ant = sumaminimos

        #Distancias euclidianas dinamicas
        for j in range(ncentros):
            for i in range(len(filas)):
                eu[j].append(math.sqrt((pow((an[i]-centro[j][0]),2))+(pow((pn[i]-centro[j][1]),2))))

        #Obetener el minimo de las distancias euclidianas
        for i in range(len(filas)):
            for j in range(ncentros):
                aux.append(eu[j][i])
            minimos.append(min(aux))
            del aux[:]

        #Obtener las listas binarias de las distancas euclidianas
        for i in range(ncentros):
            for j in range(len(filas)):
                if minimos[j] == eu[i][j]:
                    bins[i].append(1)
                else:
                    bins[i].append(0)

        #Suma de minimos
        sumaminimos = sumalista(minimos)
        print(sumaminimos)

        #Obtener los nuevos centroides
        for i in range(ncentros):
                centron[i].append(float(centro[i][0] + sumprod(an,bins[i],len(filas)))/float((sumalista(bins[i])+1)))
                centron[i].append(float(centro[i][1] + sumprod(pn,bins[i],len(filas)))/float((sumalista(bins[i])+1)))

        #Asignamos los nuevos centros a los centros actuales
        for i in range(ncentros):
            centro[i] = centron[i]

        #Condicion del ciclo
        if iteraciones >=1:
            if(sumaminimos >= ant):
                print("Mejor suma de minimos con "+str(ncentros)+" centros es "+str(ant))
                print("Coordenadas : "+str(centron))
                auxant = ant
                break


        #Vaciamos las listas
        del eu[:]
        eu = [[]for _ in range(ncentros)]
        del minimos[:]
        del bins[:]
        bins = [[]for _ in range(ncentros)]
        del centron[:]
        centron = [[]for _ in range(ncentros)]


    if(auxant >= antminimos or ncentros==8):
        print(str(auxant)+" >= "+str(antminimos))
        break

    ncentros = ncentros + 1
    del eu[:]
    eu = [[]for _ in range(ncentros)]
    del minimos[:]
    del bins[:]
    bins = [[]for _ in range(ncentros)]
    del centron[:]
    centron = [[]for _ in range(ncentros)]
    del centro[:]
    centro = [[0,0],[1,1],[0.5,0.5],[0.3,0.3],[0.4,0.4],[0.7,0.7],[0.2,0.2],[0.5,0.8],[0.9,0.9],[0.85,0.20]]

print(str(ncentros)+" fue lo optimo con una suma de minimos de "+str(ant))
print("Fin del programa")
