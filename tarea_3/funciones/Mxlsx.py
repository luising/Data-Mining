from openpyxl import load_workbook
from itertools import groupby
from operator import itemgetter


class ReadLog:
    """docstring for ."""

    def __init__(self, archivo):
        self.a = archivo
        self.wb = load_workbook(archivo)
        self.ws = self.wb.active
        filas = [[cell.value for cell in row] for row in self.ws.iter_rows()]
        self.titulo = filas[0]
        del filas[0]
        self.filas = filas

    def ordenar(self, primer, segundo):
        # parametros son las columnas que se van ordenar
        # Ordenamos las filas en función de el primero y en función del segundo.
        self.filas.sort(key=itemgetter(primer, segundo))

    def agrupar(self, columna):
        # Agrupamos en función de la columna.
        self.grupos = {k: list(v) for k, v in groupby(self.filas, itemgetter(columna))}

    def obtener_filas(self):
        return self.filas
