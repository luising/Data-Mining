from openpyxl import load_workbook
import numpy as np
import os.path as path

wbarch = load_workbook('crim.xlsx')
wsarch = wbarch.active

wbbdcrimen = load_workbook('Datos.xlsx')
wsbdcrimen = wbbdcrimen.active

#columna nivel  crimen
colJ = wsbdcrimen['J']
#columna descripcion del crimen
colF = wsarch['F']

crimenes = []
#---obtener el matriz de crimenes---
#crimenes realizados en el distrito
def comprobarCrimen(crimenA):
    archivo = open("matriz.txt","r")
    content=archivo.read().splitlines()
    for linea in content:
        crimen=linea.split(',')
        if crimen[0] in crimenA:
            archivo.close()
            return False
    archivo.close()
    return True
def confirmarCrimenes():
        print("obtener la clasificacion de los crimenes")
        archivo = open("matriz.txt","w")
        archivo.close()
        titulo = True
        for cell in colJ:
            archivo = open("matriz.txt","r")
            crimen =  wsbdcrimen.cell(row=int(cell.coordinate.split("J")[1]), column =6).value
            if crimen != None:
                if comprobarCrimen(crimen) and cell.value!=None:
                    archivo.close()
                    if titulo:
                        titulo = False
                    else:
                        #columna nivel  crimen 10
                        archivo = open("matriz.txt","a")
                        archivo.write(crimen +","+str(cell.value)+"\n")
                        archivo.close()
            obtenerCrimenes()

#----Clasificar los crimen ----
#funcion busqueda del crimen
def BDcrimen(cadena):
    for crimen in crimenes:
        #print (crimen[0]+"---"+cadena)
        if crimen[0] in cadena:
            return int(crimen[1]);
    return -1;
def obtenerCrimenes():
    archivo = open("matriz.txt","r")
    content=archivo.read().splitlines()
    for linea in content:
        crimen=linea.split(',')
        crimenes.append([crimen[0],crimen[1]])
def clasificarCrim():
    if path.exists("matriz.txt"):
        obtenerCrimenes()
        print("lista de crimenes preparada \n clasificando crimenes de archivo.." )
        for cell in colF:
            clasf= BDcrimen(cell.value)
            if clasf > 0:
                b=cell.coordinate.split("F")
                wsarch.cell(row=int(b[1]), column =10, value=clasf)
        wbarch.save('d1.xlsx')
        print("clasificacion terimnada")
    else:
        print("lista de crimenes vacia")
        confirmarCrimenes()
        clasificarCrim()
clasificarCrim()
