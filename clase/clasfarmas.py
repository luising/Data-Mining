from openpyxl import load_workbook
import numpy as np
import os.path as path

wbsan = load_workbook('Datos.xlsx')
wssan = wbsan.active

#columna Arma
colP = wssan['P']
#columna clasificacion
colT = wssan['T']
def comprobarArma(armaA):
    archivo = open("matriz.txt","r")
    content=archivo.read().splitlines()
    for linea in content:
        arma=linea.split(',')
        if  armaA in arma[0]:
            archivo.close()
            return False
    archivo.close()
    return True
def confirmarArma():
        i =0
        print("obtener la clasificacion de los armas")
        archivo = open("matriz.txt","w")
        archivo.close()
        titulo = True
        for cell in colT:
            archivo = open("matriz.txt","r")
            Arma =  wssan.cell(row=int(cell.coordinate.split("T")[1]), column =16).value
            if Arma != None:
                if comprobarArma(Arma) and cell.value!=None:
                    archivo.close()
                    if titulo:
                        titulo = False
                    else:
                        i +=1
                        #columna nivel  crimen 10
                        archivo = open("matriz.txt","a")
                        archivo.write(Arma +","+str(cell.value)+"\n")
                        archivo.close()
confirmarArma()
