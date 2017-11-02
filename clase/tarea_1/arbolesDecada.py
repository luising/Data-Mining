from openpyxl import load_workbook
from itertools import groupby
from operator import itemgetter
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from numpy import array
from PIL import Image, ImageDraw, ImageFont

archivo = "san.xlsx"
ClasificacionArmas = "M_armas.txt"
imagen = "arbol.png"
# Cargamos el archivo y la hoja correspondiente.
wb = load_workbook(filename = archivo)
ws = wb.active
#cargar por filas
filas = [ [cell.value  for cell in row ] for row in ws.iter_rows(row_offset= 1)]
del filas[len(filas)-1]
#posiciones de arbol

rraza = [0 for x in range(2)]
rgenero = [0 for x in range(4)]
rarma = [0 for x in range(8)]
indices = [4,9,6,11,15]
a = []

#Funciones
def Raza():
    if fila [4] == fila[9]:
        a.append(0)
        rraza[0] +=1
    else:
        a.append(1)
        rraza[1] +=1
    Genero()

def Genero():
    if fila[6] == fila[11]:
        a.append(0)
    else:
        a.append(1)
    indice = int(''.join(str(x) for x in a),2)
    rgenero[indice] +=1
    Arma()

def Arma():
    if comprobarArma(fila[15]) == 0:
        a.append(0)
    else:
        a.append(1)
    indice = int(''.join(str(x) for x in a),2)
    rarma[indice] +=1

def comprobarRegistro(fila):
    for indice in indices:
        if fila [indice] == None:
            return False
    return True
def comprobarArma(Arma):
    archivom = open(ClasificacionArmas,"r")
    content = archivom.read().splitlines()
    for linea in content:
        registro = linea.split(',')
        if  registro[0] in Arma :
            archivom.close()

            tipo = int (registro[1])
            return tipo
def Arbol():
    image = Image.open(imagen)
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype("arial.ttf", 45)
    if year < ultimoYear:
        draw.text((1,20), " "+str(year-10)+"-"+str(year), font=font, fill="black")
    else:
        draw.text((1,20), " "+str(year-10)+"-"+str(ultimoYear), font=font, fill="black")
    draw.text((450,30), ""+str(datos), font=font, fill="black")

    draw.text((213,160), ""+str(rraza[0]), font=font, fill="black")
    draw.text((738,160), ""+str(rraza[1]), font=font, fill="black")

    draw.text((73,242), ""+str(rgenero[0]), font=font, fill="black")
    draw.text((346,242), ""+str(rgenero[1]), font=font, fill="black")
    draw.text((602,242), ""+str(rgenero[2]), font=font, fill="black")
    draw.text((872,242), ""+str(rgenero[3]), font=font, fill="black")

    draw.text((10,326), ""+str(rarma[0]), font=font, fill="black")
    draw.text((138,326), ""+str(rarma[1]), font=font, fill="black")
    draw.text((280,326), ""+str(rarma[2]), font=font, fill="black")
    draw.text((408,326), ""+str(rarma[3]), font=font, fill="black")
    draw.text((536,326), ""+str(rarma[4]), font=font, fill="black")
    draw.text((663,326), ""+str(rarma[5]), font=font, fill="black")
    draw.text((804,326), ""+str(rarma[6]), font=font, fill="black")
    draw.text((934,326), ""+str(rarma[7]), font=font, fill="black")
    path = "arbolxDecada\\"
    image.save(path + str (i) +"arbol.png")


#Ciclo principal
datos =0
i = 1
year = 1859
ultimoYear = filas[len(filas)-1][2]
inDecada = True
while(inDecada):
    for fila in  filas:
        if fila[2] > (year-10) and fila[2] < year and comprobarRegistro(fila):
                datos += 1
                Raza()
        a = []
    Arbol()
    if year > ultimoYear:inDecada = False
    year += 10
    i += 1
