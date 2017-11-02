from openpyxl import load_workbook
from itertools import groupby
from operator import itemgetter
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
import calendar

# Cargamos el libro y la hoja correspondiente.
wb = load_workbook(filename='d1.xlsx')
ws = wb.active
# Obtenemos las filas iterando sobre la hoja.
filas = [[cell.value  for cell in row ] for row in ws.iter_rows(row_offset= 1)]
del filas[len(filas) - 1]
for fila in filas:
    if type(fila[0]) == str:
         fila[0] = datetime.strptime(fila[0], '%m/%d/%y %H:%M')
    else:
         fila[0] = fila[0].replace(day = fila[0].month, month = fila[0].day)
    if type(fila[2]) != int:
         fila[2] = int(fila[2])

# Ordenamos las filas en función de la tercera columna primero (distrito) y en función de la primera después (fechas).
filas.sort(key=itemgetter(2, 0))

# Agrupamos en función de la tercera columna (distrito).
grupos = {k: list(v) for k, v in groupby(filas,  itemgetter(2))}

for fila in filas:
    fila[0] = datetime.strptime(fila[0].strftime('%m/%d/%y'), '%m/%d/%y')
plt.title("sacramento")
plt.xlabel("mes de enero ")
plt.ylabel("indice de crimen ")

plt.xticks(np.arange(1, 32, 1))
for distrito, filas in grupos.items():
    print("--------------------- distrito " + str(distrito) + "----------------")
    lista = [fila[9] for fila in filas]
    promedios = [[sum(lista[:i + 1]) / float(i + 1), filas[i][0]] for i in range(len(lista) - 1)]
    gruposxf = {k: list(v) for k, v in groupby(promedios, itemgetter(1))}
    gruposxf = sorted(gruposxf.items(), key=itemgetter(0))
    i = 0
    promedioxmes = []
    for dia, promedios in gruposxf:
        i += 1
        print(str(i) +" dia = "+ str(dia) + " promedios =" + str(len(promedios)))
        promediosDia = [fila[0] for fila in promedios]
        promedioxdia = sum(promediosDia)/float(len(promediosDia))
        promedioxmes.append(promedioxdia)
         #[sum(promediosDia[:i+1])/float(i+1) for i in range(len(promediosDia)-1)]
    print ("tamaño de promedio " + str(len(promedioxmes)))
    plt.plot(promedioxmes ,label = "distrito"+ str(distrito))
plt.legend()
plt.show()
