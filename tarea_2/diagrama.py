import webbrowser
import random
# Librerias
from openpyxl import load_workbook
# Importacion de log y de la hoja
wb = load_workbook(filename="metro.xlsx")
ws = wb.active


def comprobar(a, ind):
    try:
        ind.index(a)
        return False
    except:
        return True


def comprobarL(a, ind):
    for i in ind:
        if a == i[1]:
            return False
    return True


def obtenerIndices(camino):
    Cind = []
    global indice
    for c in camino:
        for f in indice:
            if c == f[1]:
                Cind.append(f[0])
                break
    return Cind


def comprobarXorSplit(mat, p):
    cont = 0
    for a in mat:
        if p == a[0]:
            cont += 1
    return cont > 1


def comprobarXorJoin(mat, p):
    cont = 0
    for a in mat:
        if p == a[1]:
            cont += 1
    return cont > 1


# Obtener las filas del log en una matriz
filas = [[cell.value for cell in row] for row in ws.iter_rows(row_offset=1)]
del filas[len(filas) - 1]
print("Total de filas = " + str(len(filas)))
# Variables
casos = []
indice = []
asignado = 0

ult = filas[len(filas) - 1][0]
for i in range(1, ult + 1):
    temp = []
    for f in filas:
        if (f[0] == i):
            temp.append(f[1])
    casos.append(tuple(temp))
m = max(len(s) for s in casos)
r = [(a, s) for a, s in enumerate(casos) if len(s) == m]
print("caminos ", casos)
temp = [r[0][1]]
for a, g in enumerate(casos):
    if a == r[0][0]: continue
    else:
        temp.append(g)
casos = temp
for c in casos:
    for a in c:
        if comprobarL(a, indice):
            indice.append([asignado, a])
            asignado += 1

caminos = []
for i in range(1, ult + 1):
    temp = []
    for f in filas:
        if (f[0] == i):
            temp.append(f[1])
    caminos.append(tuple(obtenerIndices(temp)))

grupos = []
# for camino[i] in caminos:
entradaActividad = "{ \"nodeKeyProperty\": \"id\",\n\t\"nodeDataArray\": ["
entradaCamino = "\n\t],\n\t\"linkDataArray\": ["
finCamino = "\n\t]\n}"
punto = []


print(caminos)
actividades = ""
conexiones = ""
for i in range(len(caminos)):
    for a in range(len(caminos[i]) - 1):
        if comprobar((caminos[i][a], caminos[i][a + 1]), punto):
            punto.append((caminos[i][a], caminos[i][a + 1]))
            mensaje = "\"Xor Split \"" if comprobarXorSplit(punto, caminos[i][a]) else "\"\""
            mensaje = "\"Xor Join \"" if comprobarXorJoin(punto, caminos[i][a + 1]) else mensaje
            conexiones += "\n\t\t{ \"from\":" + str(caminos[i][a]) + ", \"to\": " + str(caminos[i][a + 1]) + ", \"text\":" + mensaje + " },"
conexiones = conexiones[:len(conexiones) - 1]
# \"up or timer\", \"curviness\": -20
coordenadas = [(120, 120)]
puntoB = 100
for a, c in enumerate(punto):
    if(a == 0): continue
    print(c)
    key = c[0]
    if key >= a:
        coordenadas.append((coordenadas[a - 1][0] + 300, coordenadas[a - 1][1]))
    else:
        coordenadas.append((coordenadas[key][0], coordenadas[key][1] + puntoB))


posicion = 0
for i in range(len(indice) - 1):
    x, y = coordenadas[posicion]
    actividades += "\n\t\t{ \"id\": " + str(i) + ", \"loc\": \"" + str(x) + " " + str(y) + "\", \"text\": \"" + str(indice[i][1]) + "\" },"
    posicion += 1
x, y = coordenadas[posicion + 1]
ultimo = len(indice) - 1
actividades += "\n\t\t{ \"id\": " + str(ultimo) + ", \"loc\": \"" + str(x) + " " + str(y) + "\", \"text\": \"" + str(indice[ultimo][1]) + "\" }"

# print(coordenadas)
archivo = open("p.txt", "w")
archivo.write(entradaActividad + actividades + entradaCamino + conexiones + finCamino)
archivo.close()
url = "t.html "
print("casos\t\tcaminos")
# webbrowser.open_new(url, "sadfasfa")
